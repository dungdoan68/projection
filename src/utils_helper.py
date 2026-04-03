import glob
import os.path
import shutil as sh
import pandas as pd
import time

import sys
sys.path.append(os.path.realpath("src"))
import map_SI_ILP2026
from data import data_source
import xlwings as xw
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
LabelName = 'INTERNAL v2'
LabelId = '5a27cab5-59e6-424d-abbb-227cd06504a3'

def get_Input_Value(cell, value, fileName, sheetName):
    wb = xw.Book(fileName)
    app = wb.app
    # # wb = xw.Book() # To create a new workbook
    sheet = wb.sheets[sheetName]
    sheet.range(cell).value = value
    get_wait_for_excel(wb, 0.1)


def get_Input_Value(wb, sheetName, cell, value):
    # wb = xw.books.open(Data_Source.SOURCE_FILE_ILP_2026_PUBLIC_SI)
    xw.books.active
    sheet = wb.sheets[sheetName]
    sheet.range(cell).value = value
    get_wait_for_excel(wb, 0.1)


def get_Create_Tcs(fileSource, list_fileDest):
    for item in list_fileDest:
        tc = 'Resource/UAT_Output/' + str(item[0]) + '.xlsb'
        sh.copy(fileSource, tc)
        print(f"new file '{tc}' is created")


def get_Create_Tcs_MUT(fileSource, list_fileDest):
    for item in list_fileDest:
        full = 'Resource/MUT_Output/' + 'ILPF0' + str(item[0]) + '.xlsx'
        mustPay = 'Resource/MUT_Output/' + 'ILPC0' + str(item[0]) + '.xlsx'
        sh.copy(fileSource, full)
        sh.copy(fileSource, mustPay)
        print(f"new file '{full, mustPay}' is created")


def get_Copy_New_TC_File(fileSource, fileDest):
    sh.copy(fileSource, fileDest)
    print(f"new file '{fileDest}' is created")
    return fileDest


def get_Copy_Paste_Value(source_sheet, get_Key_Values, target_sheet):
    for key in get_Key_Values:
        start_time = time.perf_counter()
        values_to_copy = source_sheet.range(get_Key_Values[key][0]).value
        # xw.Range('D1').end('down').delete()
        target_sheet.range(get_Key_Values[key][1]).value = values_to_copy
        end_time = time.perf_counter()
        elapsed_time = end_time - start_time
        print(
            "Copy " f'{key}' " from "f'{get_Key_Values[key][0]} {source_sheet}'" to " f" to {target_sheet} is OK TIME {elapsed_time}")


def get_Copy_Paste_Value_Hidden_Cells(source_sheet, get_Key_Values, target_sheet):
    for key in get_Key_Values:
        start_time = time.perf_counter()
        values_to_copy = source_sheet.range(get_Key_Values[key][0]).value
        target_sheet.range(get_Key_Values[key][1]).value = values_to_copy
        target_sheet.cells.last_cell.row
        get_Remove_Redundant_Col(target_sheet, map_SI_ILP2026.get_Remove_Redundant)
        end_time = time.perf_counter()
        elapsed_time = end_time - start_time
        print(
            "Copy " f'{key}' " from "f'{get_Key_Values[key][0]} {source_sheet}'" to " f'{target_sheet}' "is OK TIME {elapsed_time}")


def get_Copy_Paste_MUT_Value(source_sheet, sheet_Copy, cell_Copy, cell_Paste, target_sheet):
    start_time = time.perf_counter()
    values_to_copy = source_sheet.range(sheet_Copy.range(cell_Copy).value).value
    target_sheet.range(cell_Paste).value = values_to_copy
    # target_sheet.cells.last_cell.row
    get_Remove_Redundant_Col(target_sheet, map_SI_ILP2026.get_Remove_Redundant)
    end_time = time.perf_counter()
    elapsed_time = end_time - start_time
    print(
        "Copy " f'{cell_Copy}' " from "f'{sheet_Copy.range(cell_Copy)} {source_sheet}'" to " f'{target_sheet}' "is OK TIME {elapsed_time}")


def get_Copy_Paste_Fund(wb, fund_list, source_value, target_value):
    for fund in fund_list:
        elapsed_time = 0
        start_time = time.perf_counter()
        source_sheet = wb.sheets[fund_list[fund][0]]
        values_to_copy = source_sheet.range(source_value).options(ndim=2).value
        source_sheet.range(target_value).value = values_to_copy
        # get_wait_for_excel(wb, 0.1)
        end_time = time.perf_counter()
        elapsed_time = end_time - start_time
        print(
            "Copy "f'{fund}'" from "f'{source_value}'" to " f'{target_value}'f" in {source_sheet} is OK TIME {elapsed_time}")


def get_List_Tcs_(fileName, sheetName):
    workbook = load_workbook(fileName)
    sheet = workbook[sheetName]
    data = []
    for row in sheet.iter_rows(2, values_only=True):
        data.append(list(row))
    print(data)
    return data


def get_List_Tcs():
    workbook = load_workbook(data_source.TESTCASE_FILE)
    sheet = workbook['TC']
    data = []
    for row in sheet.iter_rows(2, values_only=True):
        data.append(list(row))
    print(data)
    return data


def get_Cell_Value(wb, sheet, cell):
    get_wait_for_excel(wb, 1)
    sheet_Value = wb.sheets(sheet)
    return sheet_Value.range(cell).options(ndim=2).value


def get_wait_for_excel(wb, time_wait):
    app = wb.app
    app.calculation = 'automatic'
    while app.api.CalculationState != xw.constants.CalculationState.xlDone:
        time.sleep(time_wait)
        print(time_wait)  # Sleep for a short duration to prevent blocking
    print("Excel calculation is complete.")
    try:
        app = wb.app
    except Exception as e:
        print(f"An error occurred: {e}")


def get_Copy_Low_Fund(wb, fund_list, source, target):
    for fund in fund_list:
        sheetName = fund_list[fund][0]
        get_Copy_Paste_Fund(wb, sheetName, source, target)


def get_Copy_Paste_From_TC_Sheet_To_SI_File(source_sheet, source_value, target_value, target_sheet):
    start_time = time.perf_counter()
    values_to_copy = source_sheet.range(source_value).options(ndim=2).value
    target_sheet.range(target_value).value = values_to_copy
    end_time = time.perf_counter()
    elapsed_time = end_time - start_time


def get_Delete_Column(file):
    wb = xw.Book(file)
    ws = wb.sheets['BaseAndRider']
    col_Index = 200
    while col_Index > 43:
        # for col_Index in range(200, 45):
        col_Letter = get_column_letter(col_Index)
        col = col_Letter + ':' + col_Letter
        col_range = ws.range((2, col_Index), (66, col_Index))
        col_Values = col_range.value
        # col_Values = ws.range(col).value
        numeric_Values = [
            v for v in col_Values if isinstance(v, (int, int))
        ]
        total = sum(numeric_Values)
        # total = sum(col_Values)
        print(total)
        if total == 0.0:
            ws.range(col).api.Delete()
            print(f'{col} was deleted')
        col_Index = col_Index - 1


def get_Remove_Column_Xw(ws):
    col_Index = 230
    while col_Index > 43:
        col_Letter = get_column_letter(col_Index)
        col = col_Letter + ':' + col_Letter
        col_range = ws.range(f'{col_Letter}2:{col_Letter}66')
        df = col_range.options(pd.DataFrame, index=False).value
        col_list = df.sum().tolist()
        if col_list[0] == 0.0:
            ws.range(col).delete()
            print(f'{col} was deleted')
        col_Index = col_Index - 1


def get_Remove_Column_Pandas(file, sheetName):
    df = pd.read_excel(file, sheetName)
    col_List = df.columns.to_numpy().tolist()
    print(col_List)
    col_Index = df.shape[1] - 1  #return max columns in file
    print(col_Index)
    # first_column_df = df.iloc[:, [0]]
    while col_Index > 43:
        col_Name = df.columns[col_Index]
        total = df[col_Name].sum()
        if total == 0:
            # ws.range(col).delete()
            df.drop(col_Name, axis=1, inplace=True)
            # del df[col_Name]
            print(f'{col_Name} was deleted')
        col_Index = col_Index - 1
    df.to_excel(file, sheet_name=sheetName, index=False)
    wb = xw.Book(file)
    label_info = wb.api.SensitivityLabel.CreateLabelInfo()
    label_info.AssignmentMethod = 2  # e.g., 2 for 'manual'
    label_info.LabelId = LabelId
    label_info.LabelName = LabelName
    wb.api.SensitivityLabel.SetLabel(label_info, label_info)
    wb.save()
    wb.close()


# get_Remove_Column_Pandas('Resource/MUT_Output/ILPC02078.xlsx', 'BaseAndRider')


def get_Remove_Redundant_Col(target_sheet, keys_values):
    for key in keys_values:
        target_sheet.range(keys_values[key][0], keys_values[key][1]).api.Delete()


def get_Clear_Content(ws, col):
    time.sleep(1)
    ws.range(col).clear_contents()


def get_Latest_File(path, str="*.pdf"):
    st = time.perf_counter()
    file_Path = os.path.join(path, str)
    print("RUNTIME GET SOURCE " + f'{time.perf_counter() - st}')
    st2 = time.perf_counter()
    files = sorted(glob.iglob(file_Path), key=os.path.getctime, reverse=True)
    print(files[0] + f'{time.perf_counter() - st2}')
    # print(os.path.basename(files[0]))
    return files[0]


def get_file_name(directory_path):
    file = get_Latest_File(directory_path)
    latest_file_name = os.path.basename(file).split(".pdf")[0].split("ILPF0")[1]
    print(latest_file_name)
    return latest_file_name


def get_Open_Excel_File(source_path):
    try:
        source_wb = xw.Book(source_path)
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        print('Open the workbooks')
    return source_wb


def get_Last_Row(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name) #engine='pyxlsb' use this one when need to open xlsb file
    last_row_index = len(df)+2
    print(last_row_index)
    return last_row_index

